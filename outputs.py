import datas
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


thin_borders = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def autoresize_columns(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value * 1.4
    ws.column_dimensions['A'].width = 20


def create_workbook(sheetname="Reference Ranges"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference Ranges"
    return wb, ws


def center_and_add_borders(ws, base_row, column, width=1):
    current_cell = get_column_letter(column)+str(base_row)
    ws[current_cell].alignment = Alignment(horizontal="center")

    if width > 1:
        range_string = get_column_letter(
            column)+str(base_row)+":"+get_column_letter(column+(width-1))+str(base_row)
        style_range(ws, range_string, border=thin_borders)
    else:
        ws[current_cell].border = thin_borders


def make_ok_cell(ws, current_cell):
    ws[current_cell].fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type='solid')
    ws[current_cell].font = Font(color='006100')


def make_flagged_cell(ws, current_cell):
    ws[current_cell].fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type='solid')
    ws[current_cell].font = Font(color='9C0006')


def make_warning_cell(ws, current_cell):
    ws[current_cell].fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type='solid')
    ws[current_cell].font = Font(color='9C5700')


def save_workbook(wb, filename="ranges_output.xlsx"):
    wb.save(filename)


def write_tissue_row(ws, flags_list, base_row):
    column = 1
    for i in range(1, 35):
        current_cell = get_column_letter(column)+str(base_row)
        current_flag = flags_list[i]
        ws[current_cell] = current_flag
        if current_flag == "OK":
            make_ok_cell(ws, current_cell)
        elif current_flag == "Flagged":
            make_flagged_cell(ws, current_cell)
        else:
            make_warning_cell(ws, current_cell)
        center_and_add_borders(ws, base_row, column)
        column += 1


def write_serum_row(ws, flags_list, base_row):
    column = 1
    for i in range(35, 47):
        current_cell = get_column_letter(column)+str(base_row)
        current_flag = flags_list[i]
        ws[current_cell] = current_flag
        if current_flag == "OK":
            make_ok_cell(ws, current_cell)
        elif current_flag == "Flagged":
            make_flagged_cell(ws, current_cell)
        else:
            make_warning_cell(ws, current_cell)
        center_and_add_borders(ws, base_row, column)
        column += 1


def write_serum_headers(ws, base_row):
    column = 1

    ws[get_column_letter(column)+str(base_row)] = "Copper Serum (tcus)"
    ws[get_column_letter(column)+str(base_row)
       ].alignment = Alignment(horizontal="center")
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Blood Lead (tpbb)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Blood Se (Tsemsb)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Serum Se (tsems)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    icpse_width = 7

    merge_range = get_column_letter(column)+str(base_row)+':' + \
        get_column_letter(column+icpse_width-1)+str(base_row)
    ws.merge_cells(merge_range)

    ws[get_column_letter(column)+str(base_row)] = "Mineral Panel (icpse)"
    center_and_add_borders(ws, base_row, column, width=icpse_width)

    column += icpse_width

    ws[get_column_letter(column)+str(base_row)] = "Serum Zn (tzns)"
    center_and_add_borders(ws, base_row, column)


def write_tissue_headers(ws, base_row):
    hmsc_width = 19
    merge_range = 'A'+str(base_row)+':' + \
        get_column_letter(hmsc_width-1)+str(base_row)
    ws.merge_cells(merge_range)

    column = 1

    ws[get_column_letter(column)+str(base_row)] = "Heavy Metals Panel (hmsc)"
    ws['A'+str(base_row)].alignment = Alignment(horizontal="center")
    center_and_add_borders(ws, base_row, column, width=hmsc_width)

    column = hmsc_width

    ws[get_column_letter(column)+str(base_row)] = "Copper ICP (tcu)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Copper Paired (tcup)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Lead ICP (tpb)"
    center_and_add_borders(ws, base_row, column)

    column += 1

    ws[get_column_letter(column)+str(base_row)] = "Se tissue (tseft)"
    center_and_add_borders(ws, base_row, column)

    column += 1
    icpti_width = 12

    merge_range = get_column_letter(column)+str(base_row)+':' + \
        get_column_letter(column+icpti_width-1)+str(base_row)
    ws.merge_cells(merge_range)

    ws[get_column_letter(column)+str(base_row)] = "Mineral Panel (icpti)"
    ws[get_column_letter(column)+str(base_row)
       ].alignment = Alignment(horizontal="center")
    center_and_add_borders(ws, base_row, column, width=icpti_width)


def write_tissue_element_names(ws, base_row, order):
    for column in range(1, 35):
        ws[get_column_letter(column)+str(base_row)] = order[column].title()
        center_and_add_borders(ws, base_row, column)


def write_serum_element_names(ws, base_row, order):
    i = 1
    for column in range(35, 47):
        ws[get_column_letter(i)+str(base_row)] = order[column].title()
        center_and_add_borders(ws, base_row, i)
        i += 1


def main():

    flags_list = [' ', 'OK', ' ', ' ', 'No Range', 'Flagged', 'OK', 'OK', 'OK', 'OK', 'No Range', 'OK', ' ', 'Flagged', ' ', 'No Range', ' ', ' ', 'OK', 'OK', 'No Range', 'OK', 'OK', 'OK', 'OK', 'OK', 'OK',
                  'OK', 'OK', 'OK', ' ', 'OK', 'OK', 'OK', 'OK', 'No Range', 'No Range', 'No Range', 'OK', 'OK', 'No Range', 'No Range', 'No Range', 'No Range', 'OK', 'Flagged', 'No Range']

    test_type = "Flag LOW"

    order = datas.get_elements_order()

    output_file = "ranges_output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reference Ranges"

    title_font = Font(size=16, bold=True)
    species_font = Font(size=14, bold=True)
    tissue_font = Font(size=12, bold=True)

    ws.merge_cells("A1:G1")
    ws['A1'] = "LabVantage - Checking Status of Reference Ranges"
    ws['A1'].font = title_font

    base_row = 3

    ranges = datas.load_data()

    for species in ranges:
        ws['A'+str(base_row)] = species
        ws['A'+str(base_row)].font = species_font

        base_row += 1

        for type in ranges[species]:
            ws['A'+str(base_row)] = type.title()
            ws['A'+str(base_row)].font = tissue_font
            base_row += 1

            tissue_headers = False
            serum_headers = False
            if type in ["liver", "kidney"]:
                if not tissue_headers:
                    write_tissue_headers(ws, base_row)
                    base_row += 1
                    write_tissue_element_names(ws, base_row, order)
                    tissue_headers = True
                    base_row += 1
                write_tissue_row(ws, flags_list, base_row)
                base_row += 1
            else:
                if not serum_headers:
                    write_serum_headers(ws, base_row)
                    base_row += 1
                    write_serum_element_names(ws, base_row, order)
                    serum_headers = True
                    base_row += 1
                write_serum_row(ws, flags_list, base_row)
                base_row += 1

            base_row += 2

    for i, flag in enumerate(flags_list):
        column = get_column_letter(i+1)
        ws[column + str(base_row)] = order[i]
        ws[column + str(base_row+1)] = flag

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value * 1.4
    ws.column_dimensions['A'].width = 20

    wb.save(output_file)

    # TODO: Thursday ================
    # Clean up/comment function to change specifications.
    # Set up entering data into data entry screen
    # Output organizing
    # Loop it all!

    # Shut out the lights and turn the heat down on your way out.
    print("done")


if __name__ == "__main__":
    main()
