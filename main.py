import time
import random
import datas
import outputs
from lv import *

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options


def main():

    order = datas.get_elements_order()

    excel_filename = "xlfile.xlsx"
    xl_sheetname = "Export Worksheet"

    ranges = {}

    ranges = datas.load_data(filename=excel_filename,
                             sheetname=xl_sheetname, species_dict=ranges, flags=True)

    print("Data Loaded OK")

    # Set up the web driver using Chrome since LV8 only really works with Chrome (boo).
    # Make this a global variable to make code less ornery.

    driver = Chrome()
    driver.maximize_window()

    # Set the default wait time of 10 seconds for an element to load before the script bails.
    driver.implicitly_wait(3)

    # Open LabVantage login page and make sure it exists based on the page title.
    driver.get("http://sapphire.lsd.uoguelph.ca:8080/labservices/logon.jsp")
    assert "LabVantage Logon" in driver.title

    # Call the login function. See lv.py for clarification.
    # Result should be a successful login to LV8.
    login(driver)

    # # # # # # # # # # # # # # # # # # # # # # # # # #
    # User should now be logged in, interaction with LabVantage goes below
    # # # # # # # # # # # # # # # # # # # # # # # # # #

    # Load the submission then add a new specification to it.
    bring_up_submission(driver, "18-074980")

    wb, ws = outputs.create_workbook()

    title_font = Font(size=20, bold=True)
    species_font = Font(size=16, bold=True)
    type_font = Font(size=14, bold=True)

    ws.merge_cells("A1:G1")
    ws['A1'] = "LabVantage - Checking Status of Reference Ranges"
    ws['A1'].font = title_font

    ws['A2'] = "Three rows of tests for each sample type - Below, Within Ranges, Above Range."

    base_row = 4

    for species in ranges:
        print("Beginning", species)
        ws['A'+str(base_row)] = species
        ws['A'+str(base_row)].font = species_font

        base_row += 1

        for type in ranges[species]:
            print("\tProcessing", type)
            ws['A'+str(base_row)] = type
            ws['A'+str(base_row)].font = type_font

            base_row += 1

            # Add the new specification
            select_top_sample(driver)

            element = random.choice(list(ranges[species][type]))
            spec_version_id = ranges[species][type][element]["spec_version_id"]
            clear_specifications_and_add(
                driver, species, type, spec_version_id)

            # Program gets stuck here a lot. Loop until it works!!
            in_data_entry = False
            while not in_data_entry:
                main_window = driver.current_window_handle
                # First, try to enter the data entry screen.
                try:
                    enter_data_entry(driver)
                    in_data_entry = True
                # If this fails, try again.
                except:
                    # If this try succeeds, great, enter data entry.
                    # If not, try closing the specs window again first.
                    try:
                        print("\t\tRetrying to exit specs window.")
                        driver.switch_to.window(main_window)
                        driver.switch_to.default_content()
                        dlg_frame = driver.find_element_by_tag_name("iframe")
                        driver.switch_to.frame(dlg_frame[3])
                        exit_specifications_window(driver)
                    except:
                        pass

            tissue_headers = False
            serum_headers = False

            test_types = ["flag_low_value", "flag_ok_value", "flag_high_value"]
            for test_type in test_types:

                input_string = datas.get_input_string(
                    ranges, species, type, test_type)
                clear_inputs_and_paste_new(driver)
                flags_list = check_data_flags(driver)

                if type in ["liver", "kidney"]:
                    if not tissue_headers:
                        outputs.write_tissue_headers(ws, base_row)
                        base_row += 1
                        outputs.write_tissue_element_names(ws, base_row, order)
                        tissue_headers = True
                        base_row += 1
                    outputs.write_tissue_row(
                        ws, flags_list, base_row, input_string)
                    base_row += 1
                else:
                    if not serum_headers:
                        outputs.write_serum_headers(ws, base_row)
                        base_row += 1
                        outputs.write_serum_element_names(ws, base_row, order)
                        serum_headers = True
                        base_row += 1
                    outputs.write_serum_row(
                        ws, flags_list, base_row, input_string)
                    base_row += 1

            exit_data_entry(driver)
            outputs.autoresize_columns(ws)
            outputs.save_workbook(wb)
            base_row += 2

    outputs.autoresize_columns(ws)

    outputs.save_workbook(wb)

    time.sleep(5)

    # # # # # #
    # Interactions in LabVantage should all be above this, teardown only past this point.
    # # # # # #

    # Teardown.

    logout(driver)
    driver.close()
    driver.quit()

    # Shut out the lights and turn the heat down on your way out.
    print("done")


if __name__ == "__main__":
    main()
